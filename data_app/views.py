from openpyxl import load_workbook
from django.shortcuts import render, redirect
from .forms import ExcelFileForm
from .models import ExcelFile

def upload_file(request):
    if request.method == 'POST':
        form = ExcelFileForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('view_data')
    else:
        form = ExcelFileForm()
    return render(request, 'data_app/upload.html', {'form': form})

def view_data(request):
    latest_file = ExcelFile.objects.latest('uploaded_at')
    file_path = latest_file.file.path

    # Load the workbook and the first sheet
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    # Extract data from the sheet and convert it to HTML
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Create an HTML table from the data
    html_data = "<table border='1'>"
    for row in data:
        html_data += "<tr>"
        for cell in row:
            html_data += f"<td>{cell}</td>"
        html_data += "</tr>"
    html_data += "</table>"

    return render(request, 'data_app/view_data.html', {'data': html_data})
